import time
import openpyxl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import random
import re
from lxml import etree

#options = webdriver.ChromeOptions()
#options.add_experimental_option('prefs', {'profile.managed_default_content_settings.images': 2})
#不加载图片
chromeDriverPath = r'.\tools\chromedriver.exe'
browser = webdriver.Chrome(executable_path=chromeDriverPath)
browser.maximize_window()# 窗口最大化 <br>>>> #模拟鼠标悬浮
wait =WebDriverWait(browser,50)#设置等待时间
url = 'http://www.shanyaoo.com/'
data_list= []#设置全局变量用来存储数据
username = ""#用户名
password = ""#登录密码
keyword ="中药"#关键词

def search():
    browser.get(url)

    browser.find_element_by_xpath('//input[@placeholder="用户名"]').send_keys(username)
    browser.find_element_by_xpath('//input[@placeholder="登录密码"]').send_keys(password)
    browser.find_element_by_xpath('//button[@class="loginBtn"]').click()
    try:
        input = wait.until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "input.search-input"))
        )  #等到搜索框加载出来
        submit = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='searchForm ']/div/button[1]"))
        )#等到搜索按钮可以被点击
        input[0].send_keys(keyword)#向搜索框内输入关键词
        submit.click()#点击
        total = wait.until(
            EC.presence_of_all_elements_located((By.XPATH, './/div[@class="pagebar"]//span'))
        )
        for tl in total:
            total = re.findall(r"\d+\.?\d*", tl.text)[0]
            print("总共页数："+total)
        #记录一下总页码,等到总页码加载出来
        # # 滑动到底部，加载出后三十个货物信息
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        wait.until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#nohide_one"))
        )
        browser.find_element_by_link_text("只看有货").click()#模拟用户点击
        time.sleep(5)#设置随机延迟
        html = browser.page_source#获取网页信息
        time.sleep(20)#设置随机延迟
        prase_html(html)#调用提取数据的函数
        #返回总页数
        return total
    except TimeoutError:
        search()

def next_page(page_number):
    try:
        # 滑动到底部
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.randint(1, 3))#设置随机延迟
        button = wait.until(
            EC.element_to_be_clickable((By.CLASS_NAME, 'nextpage'))
        )#翻页按钮
        button.click()# 翻页动作
        wait.until(
            EC.presence_of_all_elements_located((By.XPATH, "//*[@id='nohide_one']"))
        )#等到30个商品都加载出来
        # 滑动到底部，加载出后三十个货物信息
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        wait.until(
            EC.presence_of_all_elements_located((By.XPATH, "//*[@id='nohide_one']"))
        )#等到60个商品都加载出来
        wait.until(
            EC.text_to_be_present_in_element((By.CLASS_NAME, "cur"), str(page_number))
        )# 判断翻页成功,高亮的按钮数字与设置的页码一样
        browser.find_element_by_link_text("只看有货").click()#模拟用户点击
        html = browser.page_source#获取网页信息
        time.sleep(20)#设置随机延迟
        #调用提取数据的函数
        prase_html(html)
    except TimeoutError:
        return next_page(page_number)

#提取页面数据
def prase_html(html):
    html = etree.HTML(html)
    # 开始提取信息,找到ul标签下的全部li标签
    try:
        lis = browser.find_elements_by_class_name('pl-skin')
        # 遍历
        for pl in lis:
            #判断有禁售药品或下架
            if pl.get_attribute("id")=="nohide_one":
                # 药品名称
                title = pl.find_element_by_xpath('.//div[@class="p-caption"]//a').text
                # 售价
                price = pl.find_element_by_xpath('.//span[@class="price"]//em')
                # 销量
                sale = pl.find_element_by_xpath('.//span[@class="p-sale"]//em')
                # 阶梯满减或满减
                li_caption = pl.find_elements_by_xpath('.//li[@class="promt_li"]//a')
                if price.text:
                    price = price.text
                else:
                    price = "商家尚未定价"
                if sale:
                    sale = sale.text
                else:
                    sale = ""
                if li_caption:
                    # for labx in li_caption:
                    if li_caption[0].text=="阶梯满减":
                        j_promt_caption = li_caption[1].get_attribute("innerHTML")
                        m_promt_caption = li_caption[3].get_attribute("innerHTML")
                    elif li_caption[0].text=="满减":
                        m_promt_caption = li_caption[1].get_attribute("innerHTML")
                        j_promt_caption = ""
                    elif li_caption[2].text=="满减":
                        j_promt_caption = ""
                        m_promt_caption = li_caption[3].get_attribute("innerHTML")
                    else:
                        j_promt_caption = ""
                        m_promt_caption = ""
                else:
                    j_promt_caption = ""
                    m_promt_caption = ""
                data_dict = []#写入字典
                data_dict.append(title)
                data_dict.append( price)
                data_dict.append(sale)
                data_dict.append(j_promt_caption)
                data_dict.append(m_promt_caption)
                print(data_dict)
                data_list.append(data_dict)#写入全局变量
    except TimeoutError:
        prase_html(html)

#新建excel
def creatwb(wbname):
    wb=openpyxl.Workbook()
    wb.save(filename=wbname)
    print ("新建Excel："+wbname+"成功")

#保存页面数据
def write_excel(fileName,d_list):
    wb = Workbook()
    #写入表头
    dilei_head = ['药品名称','售价','销量','阶梯满减','满减']
    sheet0Name = '药品数据信息'
    sheet0 = wb.create_sheet(sheet0Name, index=0)
    for i, item in enumerate(dilei_head):
        sheet0.cell(row = 1,column=i+1,value=item.encode('utf-8'))
    #写入数据
    for i, data in enumerate(d_list):
        for j, item in enumerate(data):
            sheet0.cell(row = i+2,column=j+1,value=item.encode('utf-8'))

    wb.save(fileName+'.xlsx')
    print('excel文件写入完成')

def main():
    #初始化exlce文件
    creatwb("yp.xlsx")
    print("第", 1, "页：")
    total = int(search())
    # for i in range(2, 5):
    for i in range(2, total + 1):
        time.sleep(20)  # 设置随机延迟
        print("第", i, "页：")
        next_page(i)
    write_excel("yp",data_list)

if __name__ == "__main__":
    main()