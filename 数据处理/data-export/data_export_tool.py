import json
import os
import csv
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, timedelta
import calendar
from dateutil.relativedelta import relativedelta
import threading
import time
import schedule
import logging
import sys

# 尝试导入 tkcalendar，如果失败则尝试从资源路径导入
try:
    from tkcalendar import DateEntry
except ImportError:
    # 如果是打包环境，尝试从资源路径导入
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
        sys.path.append(os.path.join(base_dir, 'tkcalendar'))
        from tkcalendar import DateEntry
    else:
        raise

class DataExporter:
    @staticmethod
    def fetch_data_from_api(api_base_url, id_param, start_time, end_time):
        """从API接口获取数据"""
        # 完整的API地址 = 基础地址 + 固定接口路径
        api_url = f"{api_base_url}/api/dataRetrievalFile"
        
        payload = {
            "id": id_param,
            "startTime": start_time,
            "endTime": end_time
        }
        headers = {"Content-Type": "application/json"}
        
        try:
            response = requests.post(api_url, json=payload, headers=headers)
            if response.status_code == 200:
                return response.text
            else:
                logging.error(f"API请求失败: {response.status_code} - {response.text}")
                return None
        except requests.exceptions.RequestException as e:
            logging.error(f"请求错误: {e}")
            return None

    @staticmethod
    def parse_api_response(json_data):
        """解析API返回的JSON数据"""
        if isinstance(json_data, str):
            try:
                data = json.loads(json_data)
            except json.JSONDecodeError as e:
                logging.error(f"JSON解析错误: {e}")
                return None
        else:
            data = json_data
            
        meta = data.get('meta', {})
        if meta.get('code') != 200:
            logging.error(f"API返回错误: {meta.get('code')} - {meta.get('msg')}")
            return None
            
        result = data.get('data', {}).get('result', {})
        cellset = result.get('cellset', [])
        
        if not cellset:
            logging.error("未找到cellset数据")
            return None
            
        return cellset

    @staticmethod
    def export_to_csv(cellset, file_path):
        """导出为CSV文件"""
        if cellset is None:
            return False
        
        try:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                csv_writer = csv.writer(csvfile)
                
                # 表头
                if cellset:
                    header_row = cellset[0]
                    headers = [cell['value'] for cell in header_row]
                    csv_writer.writerow(headers)
                    
                    # 数据行
                    for row in cellset[1:]:
                        values = []
                        for cell in row:
                            if cell.get('type') == 'DATA_CELL':
                                properties = cell.get('properties', {})
                                if 'raw' in properties and properties['raw'] is not None:
                                    values.append(str(properties['raw']))
                                else:
                                    values.append(cell['value'])
                            else:
                                values.append(cell['value'])
                        csv_writer.writerow(values)
            
            logging.info(f"CSV文件已导出: {file_path}")
            return True
        except Exception as e:
            logging.error(f"导出CSV错误: {e}")
            return False

    @staticmethod
    def export_to_excel(cellset, file_path):
        """导出为Excel文件"""
        if cellset is None:
            return False
        
        try:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            
            if ws is None:
                logging.error("无法创建工作表")
                return False
                
            ws.title = "查询结果"
            
            # 样式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 表头
            if cellset:
                header_row = cellset[0]
                for col_idx, cell in enumerate(header_row, 1):
                    ws.cell(row=1, column=col_idx, value=cell['value'])
                    header_cell = ws.cell(row=1, column=col_idx)
                    header_cell.font = header_font
                    header_cell.fill = header_fill
                    header_cell.border = thin_border
                    header_cell.alignment = center_alignment
                
                # 数据行
                for row_idx, row in enumerate(cellset[1:], 2):
                    for col_idx, cell in enumerate(row, 1):
                        if cell.get('type') == 'DATA_CELL':
                            properties = cell.get('properties', {})
                            if 'raw' in properties and properties['raw'] is not None:
                                value = properties['raw']
                            else:
                                value = cell['value']
                        else:
                            value = cell['value']
                        
                        try:
                            if isinstance(value, str):
                                num_value = float(value.replace(',', ''))
                                ws.cell(row=row_idx, column=col_idx, value=num_value)
                            else:
                                ws.cell(row=row_idx, column=col_idx, value=value)
                        except (ValueError, TypeError):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        data_cell = ws.cell(row=row_idx, column=col_idx)
                        data_cell.border = thin_border
            
            # 调整列宽
            for i, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(i)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            logging.info(f"Excel文件已导出: {file_path}")
            return True
        except Exception as e:
            logging.error(f"导出Excel错误: {e}")
            return False

    @staticmethod
    def export_to_excel_with_pandas(cellset, file_path):
        """使用Pandas导出Excel文件"""
        if cellset is None:
            return False
        
        try:
            if cellset:
                headers = [cell['value'] for cell in cellset[0]]
                rows = []
                
                for row in cellset[1:]:
                    row_data = []
                    for cell in row:
                        if cell.get('type') == 'DATA_CELL':
                            properties = cell.get('properties', {})
                            if 'raw' in properties and properties['raw'] is not None:
                                row_data.append(properties['raw'])
                            else:
                                row_data.append(cell['value'])
                        else:
                            row_data.append(cell['value'])
                    rows.append(row_data)
                
                df = pd.DataFrame(rows)
                if headers:
                    df.columns = headers
                
                os.makedirs(os.path.dirname(file_path), exist_ok=True)
                df.to_excel(file_path, index=False)
                logging.info(f"Excel文件已导出: {file_path}")
                return True
            return False
        except Exception as e:
            logging.error(f"Pandas导出错误: {e}")
            return False

class DataExporterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("数据导出工具")
        self.root.geometry("800x700")
        self.root.resizable(True, True)
        
        # 设置日志
        self.setup_logging()
        
        # 默认API基础配置
        self.api_base_url = "http://127.0.0.1:8088/bi"
        
        # 定时任务相关变量
        self.scheduler_running = False
        self.scheduler_thread = None
        
        # 创建主框架
        self.main_frame = ttk.Frame(root, padding="20")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建输入字段
        self.create_input_fields()
        
        # 创建按钮
        self.create_buttons()
        
        # 创建定时任务控件
        self.create_scheduler_controls()
        
        # 状态标签
        self.status_label = ttk.Label(self.main_frame, text="就绪", foreground="green")
        self.status_label.grid(row=11, column=0, columnspan=3, pady=10, sticky="w")
    
    def setup_logging(self):
        """设置日志配置"""
        if getattr(sys, 'frozen', False):
            # 如果是打包的EXE环境，使用EXE所在目录
            base_dir = os.path.dirname(sys.executable)
            # 添加资源路径到系统路径
            sys.path.append(os.path.join(base_dir, 'tkcalendar'))
            sys.path.append(os.path.join(base_dir, 'openpyxl'))
        else:
            # 否则，使用脚本所在目录
            base_dir = os.path.dirname(os.path.abspath(__file__))
        
        log_file = os.path.join(base_dir, "data_export.log")
        
        # 显式检查并创建空文件（如果不存在）
        if not os.path.exists(log_file):
            with open(log_file, 'w', encoding='utf-8') as f:
                pass  # 创建空文件
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        logging.info("数据导出工具启动")
    
    def create_input_fields(self):
        """创建输入字段"""
        # API基础地址配置
        ttk.Label(self.main_frame, text="API地址:").grid(row=0, column=0, sticky="w", pady=5)
        self.api_base_url_var = tk.StringVar(value=self.api_base_url)
        self.api_base_url_entry = ttk.Entry(self.main_frame, textvariable=self.api_base_url_var, width=50)
        self.api_base_url_entry.grid(row=0, column=1, columnspan=2, sticky="we", pady=5)
        
        # ID输入
        ttk.Label(self.main_frame, text="ID:").grid(row=1, column=0, sticky="w", pady=5)
        self.id_entry = ttk.Entry(self.main_frame, width=20)
        self.id_entry.grid(row=1, column=1, sticky="w", pady=5)
        self.id_entry.insert(0, "9508013")
        
        # 时间粒度选择
        ttk.Label(self.main_frame, text="时间粒度:").grid(row=2, column=0, sticky="w", pady=5)
        self.granularity_var = tk.StringVar(value="month")
        granularity_frame = ttk.Frame(self.main_frame)
        granularity_frame.grid(row=2, column=1, sticky="w")
        
        ttk.Radiobutton(granularity_frame, text="年", variable=self.granularity_var, value="year", command=self.update_time_controls).pack(side=tk.LEFT)
        ttk.Radiobutton(granularity_frame, text="月", variable=self.granularity_var, value="month", command=self.update_time_controls).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(granularity_frame, text="日", variable=self.granularity_var, value="day", command=self.update_time_controls).pack(side=tk.LEFT)
        ttk.Radiobutton(granularity_frame, text="季度", variable=self.granularity_var, value="quarter", command=self.update_time_controls).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(granularity_frame, text="周", variable=self.granularity_var, value="week", command=self.update_time_controls).pack(side=tk.LEFT)
        
        # 时间控件容器
        self.time_controls_frame = ttk.Frame(self.main_frame)
        self.time_controls_frame.grid(row=3, column=0, columnspan=3, sticky="we", pady=5)
        
        # 初始化时间控件
        self.create_time_controls()
        
        # 备份文件夹
        ttk.Label(self.main_frame, text="备份文件夹:").grid(row=4, column=0, sticky="w", pady=5)
        self.backup_folder_var = tk.StringVar()
        self.backup_folder_entry = ttk.Entry(self.main_frame, textvariable=self.backup_folder_var, width=40)
        self.backup_folder_entry.grid(row=4, column=1, sticky="we", pady=5)
        
        # 上传文件夹
        ttk.Label(self.main_frame, text="上传文件夹:").grid(row=5, column=0, sticky="w", pady=5)
        self.upload_folder_var = tk.StringVar()
        self.upload_folder_entry = ttk.Entry(self.main_frame, textvariable=self.upload_folder_var, width=40)
        self.upload_folder_entry.grid(row=5, column=1, sticky="we", pady=5)
        
        # 文件格式选择
        ttk.Label(self.main_frame, text="文件格式:").grid(row=6, column=0, sticky="w", pady=5)
        self.format_frame = ttk.Frame(self.main_frame)
        self.format_frame.grid(row=6, column=1, sticky="w")
        
        self.format_var = tk.StringVar(value="Excel")
        ttk.Radiobutton(self.format_frame, text="Excel (.xlsx)", variable=self.format_var, value="Excel").pack(side=tk.LEFT)
        ttk.Radiobutton(self.format_frame, text="CSV (.csv)", variable=self.format_var, value="CSV").pack(side=tk.LEFT, padx=10)
        
        # 文件名格式
        ttk.Label(self.main_frame, text="文件名格式:").grid(row=7, column=0, sticky="w", pady=5)
        self.filename_format_var = tk.StringVar(value="hqmsmjz_YYYYMmm")
        self.filename_format_entry = ttk.Entry(self.main_frame, textvariable=self.filename_format_var, width=40)
        self.filename_format_entry.grid(row=7, column=1, sticky="we", pady=5)
        
        # 文件名预览
        ttk.Label(self.main_frame, text="文件名预览:").grid(row=8, column=0, sticky="w", pady=5)
        self.filename_preview_var = tk.StringVar()
        self.filename_preview_label = ttk.Label(self.main_frame, textvariable=self.filename_preview_var)
        self.filename_preview_label.grid(row=8, column=1, sticky="w", pady=5)
        
        # 更新文件名预览
        self.update_filename_preview()
    
    def create_time_controls(self):
        """创建时间控件"""
        # 清空容器
        for widget in self.time_controls_frame.winfo_children():
            widget.destroy()
        
        granularity = self.granularity_var.get()
        
        # 开始时间标签
        ttk.Label(self.time_controls_frame, text="开始时间:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        
        # 结束时间标签
        ttk.Label(self.time_controls_frame, text="结束时间:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        
        if granularity == "day":
            # 日期选择控件 - 仅用于"日"粒度
            self.start_date_var = tk.StringVar()
            self.start_date_entry = DateEntry(
                self.time_controls_frame, 
                textvariable=self.start_date_var,
                date_pattern='yyyy-mm-dd',
                width=12
            )
            self.start_date_entry.grid(row=0, column=1, sticky="w", padx=5, pady=5)
            
            self.end_date_var = tk.StringVar()
            self.end_date_entry = DateEntry(
                self.time_controls_frame, 
                textvariable=self.end_date_var,
                date_pattern='yyyy-mm-dd',
                width=12
            )
            self.end_date_entry.grid(row=1, column=1, sticky="w", padx=5, pady=5)
            
            # 设置默认时间为当前日期
            today = datetime.now()
            self.start_date_entry.set_date(today)
            self.end_date_entry.set_date(today)
            
        elif granularity == "month":
            # 为"月"粒度创建专门的年月选择器
            current_year = datetime.now().year
            current_month = datetime.now().month
            
            # 年份选项
            years = [str(y) for y in range(current_year - 10, current_year + 11)]
            # 月份选项
            months = [f"{m:02d}" for m in range(1, 13)]
            
            # 开始时间 - 年和月
            start_frame = ttk.Frame(self.time_controls_frame)
            start_frame.grid(row=0, column=1, sticky="w", padx=5, pady=5)
            
            self.start_year_var = tk.StringVar(value=str(current_year))
            ttk.Combobox(start_frame, textvariable=self.start_year_var, values=years, width=8).pack(side=tk.LEFT)
            ttk.Label(start_frame, text="年").pack(side=tk.LEFT, padx=2)
            
            self.start_month_var = tk.StringVar(value=f"{current_month:02d}")
            ttk.Combobox(start_frame, textvariable=self.start_month_var, values=months, width=5).pack(side=tk.LEFT, padx=5)
            ttk.Label(start_frame, text="月").pack(side=tk.LEFT)
            
            # 结束时间 - 年和月
            end_frame = ttk.Frame(self.time_controls_frame)
            end_frame.grid(row=1, column=1, sticky="w", padx=5, pady=5)
            
            self.end_year_var = tk.StringVar(value=str(current_year))
            ttk.Combobox(end_frame, textvariable=self.end_year_var, values=years, width=8).pack(side=tk.LEFT)
            ttk.Label(end_frame, text="年").pack(side=tk.LEFT, padx=2)
            
            self.end_month_var = tk.StringVar(value=f"{current_month:02d}")
            ttk.Combobox(end_frame, textvariable=self.end_month_var, values=months, width=5).pack(side=tk.LEFT, padx=5)
            ttk.Label(end_frame, text="月").pack(side=tk.LEFT)
            
        elif granularity == "year":
            # 年份选择控件
            current_year = datetime.now().year
            years = [str(y) for y in range(current_year - 10, current_year + 11)]
            
            self.start_year_var = tk.StringVar(value=str(current_year))
            self.start_year_combo = ttk.Combobox(self.time_controls_frame, textvariable=self.start_year_var, values=years, width=8)
            self.start_year_combo.grid(row=0, column=1, sticky="w", padx=5, pady=5)
            
            self.end_year_var = tk.StringVar(value=str(current_year))
            self.end_year_combo = ttk.Combobox(self.time_controls_frame, textvariable=self.end_year_var, values=years, width=8)
            self.end_year_combo.grid(row=1, column=1, sticky="w", padx=5, pady=5)
            
        elif granularity == "quarter":
            # 季度选择控件
            current_year = datetime.now().year
            years = [str(y) for y in range(current_year - 10, current_year + 11)]
            quarters = ["1", "2", "3", "4"]
            
            # 开始时间
            start_frame = ttk.Frame(self.time_controls_frame)
            start_frame.grid(row=0, column=1, sticky="w", padx=5, pady=5)
            
            self.start_year_var = tk.StringVar(value=str(current_year))
            ttk.Combobox(start_frame, textvariable=self.start_year_var, values=years, width=8).pack(side=tk.LEFT)
            ttk.Label(start_frame, text="季度:").pack(side=tk.LEFT, padx=5)
            self.start_quarter_var = tk.StringVar(value="1")
            ttk.Combobox(start_frame, textvariable=self.start_quarter_var, values=quarters, width=4).pack(side=tk.LEFT)
            
            # 结束时间
            end_frame = ttk.Frame(self.time_controls_frame)
            end_frame.grid(row=1, column=1, sticky="w", padx=5, pady=5)
            
            self.end_year_var = tk.StringVar(value=str(current_year))
            ttk.Combobox(end_frame, textvariable=self.end_year_var, values=years, width=8).pack(side=tk.LEFT)
            ttk.Label(end_frame, text="季度:").pack(side=tk.LEFT, padx=5)
            self.end_quarter_var = tk.StringVar(value="4")
            ttk.Combobox(end_frame, textvariable=self.end_quarter_var, values=quarters, width=4).pack(side=tk.LEFT)
            
        elif granularity == "week":
            # 周选择控件
            current_year = datetime.now().year
            years = [str(y) for y in range(current_year - 10, current_year + 11)]
            weeks = [str(w).zfill(2) for w in range(1, 53)]
            
            # 开始时间
            start_frame = ttk.Frame(self.time_controls_frame)
            start_frame.grid(row=0, column=1, sticky="w", padx=5, pady=5)
            
            self.start_year_var = tk.StringVar(value=str(current_year))
            ttk.Combobox(start_frame, textvariable=self.start_year_var, values=years, width=8).pack(side=tk.LEFT)
            ttk.Label(start_frame, text="周:").pack(side=tk.LEFT, padx=5)
            self.start_week_var = tk.StringVar(value="01")
            ttk.Combobox(start_frame, textvariable=self.start_week_var, values=weeks, width=4).pack(side=tk.LEFT)
            
            # 结束时间
            end_frame = ttk.Frame(self.time_controls_frame)
            end_frame.grid(row=1, column=1, sticky="w", padx=5, pady=5)
            
            self.end_year_var = tk.StringVar(value=str(current_year))
            ttk.Combobox(end_frame, textvariable=self.end_year_var, values=years, width=8).pack(side=tk.LEFT)
            ttk.Label(end_frame, text="周:").pack(side=tk.LEFT, padx=5)
            self.end_week_var = tk.StringVar(value="01")
            ttk.Combobox(end_frame, textvariable=self.end_week_var, values=weeks, width=4).pack(side=tk.LEFT)
    
    def update_time_controls(self):
        """更新时间控件"""
        self.create_time_controls()
    
    def create_buttons(self):
        """创建按钮"""
        # 备份文件夹浏览按钮
        backup_browse_button = ttk.Button(self.main_frame, text="浏览...", command=self.browse_backup_folder)
        backup_browse_button.grid(row=4, column=2, sticky="w", padx=5)
        
        # 上传文件夹浏览按钮
        upload_browse_button = ttk.Button(self.main_frame, text="浏览...", command=self.browse_upload_folder)
        upload_browse_button.grid(row=5, column=2, sticky="w", padx=5)
        
        # 快速日期范围按钮
        ttk.Label(self.main_frame, text="快速日期范围:").grid(row=9, column=0, sticky="w", pady=5)
        date_range_frame = ttk.Frame(self.main_frame)
        date_range_frame.grid(row=9, column=1, sticky="w", pady=5)
        
        ttk.Button(date_range_frame, text="今天", command=lambda: self.set_date_range("today")).pack(side=tk.LEFT, padx=2)
        ttk.Button(date_range_frame, text="本周", command=lambda: self.set_date_range("week")).pack(side=tk.LEFT, padx=2)
        ttk.Button(date_range_frame, text="本月", command=lambda: self.set_date_range("month")).pack(side=tk.LEFT, padx=2)
        ttk.Button(date_range_frame, text="本季度", command=lambda: self.set_date_range("quarter")).pack(side=tk.LEFT, padx=2)
        ttk.Button(date_range_frame, text="今年", command=lambda: self.set_date_range("year")).pack(side=tk.LEFT, padx=2)
        
        # 导出按钮
        export_button = ttk.Button(self.main_frame, text="导出数据", command=self.export_data)
        export_button.grid(row=10, column=0, columnspan=3, pady=10)
    
    def create_scheduler_controls(self):
        """创建定时任务控件"""
        # 定时任务标题
        ttk.Separator(self.main_frame, orient='horizontal').grid(row=11, column=0, columnspan=3, sticky="we", pady=10)
        ttk.Label(self.main_frame, text="定时任务设置", font=("Arial", 10, "bold")).grid(row=12, column=0, sticky="w", pady=5)
        
        # 定时任务框架
        scheduler_frame = ttk.Frame(self.main_frame)
        scheduler_frame.grid(row=12, column=1, columnspan=2, sticky="we", pady=5)
        
        # 定时任务类型
        ttk.Label(scheduler_frame, text="执行频率:").grid(row=0, column=0, sticky="w", padx=5)
        self.schedule_type_var = tk.StringVar(value="每天")
        schedule_type_combo = ttk.Combobox(scheduler_frame, textvariable=self.schedule_type_var, 
                                          values=["每天", "每周", "每月", "每小时"], width=10)
        schedule_type_combo.grid(row=0, column=1, sticky="w", padx=5)
        schedule_type_combo.bind("<<ComboboxSelected>>", self.update_schedule_time_controls)
        
        # 定时任务时间控件
        self.schedule_time_frame = ttk.Frame(scheduler_frame)
        self.schedule_time_frame.grid(row=0, column=2, sticky="w", padx=5)
        
        # 初始化时间控件
        self.update_schedule_time_controls()
        
        # 定时任务按钮
        self.schedule_button = ttk.Button(scheduler_frame, text="启动定时任务", command=self.toggle_scheduler)
        self.schedule_button.grid(row=0, column=3, sticky="w", padx=10)
    
    def update_schedule_time_controls(self, event=None):
        """更新定时任务时间控件"""
        # 清空容器
        for widget in self.schedule_time_frame.winfo_children():
            widget.destroy()
        
        schedule_type = self.schedule_type_var.get()
        
        if schedule_type == "每天":
            # 每天特定时间
            self.schedule_time_var = tk.StringVar(value="09:00")
            time_entry = ttk.Entry(self.schedule_time_frame, textvariable=self.schedule_time_var, width=8)
            time_entry.pack(side=tk.LEFT)
            ttk.Label(self.schedule_time_frame, text="(HH:MM)").pack(side=tk.LEFT, padx=5)
        elif schedule_type == "每周":
            # 每周特定星期几和时间
            days = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
            self.schedule_day_var = tk.StringVar(value="星期一")
            day_combo = ttk.Combobox(self.schedule_time_frame, textvariable=self.schedule_day_var, values=days, width=10)
            day_combo.pack(side=tk.LEFT)
            
            self.schedule_time_var = tk.StringVar(value="09:00")
            time_entry = ttk.Entry(self.schedule_time_frame, textvariable=self.schedule_time_var, width=8)
            time_entry.pack(side=tk.LEFT, padx=5)
            ttk.Label(self.schedule_time_frame, text="(HH:MM)").pack(side=tk.LEFT, padx=5)
        elif schedule_type == "每月":
            # 每月特定日期和时间
            self.schedule_day_var = tk.StringVar(value="1")
            day_combo = ttk.Combobox(self.schedule_time_frame, textvariable=self.schedule_day_var, 
                                    values=[str(i) for i in range(1, 32)], width=5)
            day_combo.pack(side=tk.LEFT)
            ttk.Label(self.schedule_time_frame, text="日").pack(side=tk.LEFT, padx=2)
            
            self.schedule_time_var = tk.StringVar(value="09:00")
            time_entry = ttk.Entry(self.schedule_time_frame, textvariable=self.schedule_time_var, width=8)
            time_entry.pack(side=tk.LEFT, padx=5)
            ttk.Label(self.schedule_time_frame, text="(HH:MM)").pack(side=tk.LEFT, padx=5)
        elif schedule_type == "每小时":
            # 每小时特定分钟
            ttk.Label(self.schedule_time_frame, text="每小时的").pack(side=tk.LEFT)
            self.schedule_minute_var = tk.StringVar(value="0")
            minute_combo = ttk.Combobox(self.schedule_time_frame, textvariable=self.schedule_minute_var, 
                                       values=[str(i) for i in range(0, 60)], width=5)
            minute_combo.pack(side=tk.LEFT, padx=5)
            ttk.Label(self.schedule_time_frame, text="分").pack(side=tk.LEFT)
    
    def browse_backup_folder(self):
        """浏览备份文件夹"""
        folder_path = filedialog.askdirectory(title="选择备份文件夹")
        if folder_path:
            self.backup_folder_var.set(folder_path)
            self.update_filename_preview()
    
    def browse_upload_folder(self):
        """浏览上传文件夹"""
        folder_path = filedialog.askdirectory(title="选择上传文件夹")
        if folder_path:
            self.upload_folder_var.set(folder_path)
            self.update_filename_preview()
    
    def update_filename_preview(self):
        """更新文件名预览"""
        filename = self.generate_filename()
        self.filename_preview_var.set(filename)
    
    def generate_filename(self):
        """生成文件名"""
        # 获取文件名格式
        filename_format = self.filename_format_var.get()
        
        # 获取当前日期
        now = datetime.now()
        year = now.strftime("%Y")
        month = now.strftime("%m")
        
        # 替换占位符
        filename = filename_format.replace("YYYY", year)
        filename = filename.replace("yy", year[2:])
        filename = filename.replace("mm", month)
        filename = filename.replace("MM", month)
        
        # 添加文件扩展名
        file_format = self.format_var.get().lower()
        if file_format == "excel":
            filename += ".xlsx"
        else:
            filename += ".csv"
            
        return filename
    
    def get_time_values(self):
        """获取时间值"""
        granularity = self.granularity_var.get()
        
        if granularity == "day":
            start_date = self.start_date_entry.get_date()
            end_date = self.end_date_entry.get_date()
            return start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")
            
        elif granularity == "month":
            # 使用年月选择器
            start_year = self.start_year_var.get()
            start_month = self.start_month_var.get()
            end_year = self.end_year_var.get()
            end_month = self.end_month_var.get()
            
            return f"{start_year}-{start_month}", f"{end_year}-{end_month}"
            
        elif granularity == "year":
            return self.start_year_var.get(), self.end_year_var.get()
            
        elif granularity == "quarter":
            start_quarter = self.start_quarter_var.get()
            end_quarter = self.end_quarter_var.get()
            return f"{self.start_year_var.get()}-{start_quarter}", f"{self.end_year_var.get()}-{end_quarter}"
            
        elif granularity == "week":
            start_week = self.start_week_var.get()
            end_week = self.end_week_var.get()
            return f"{self.start_year_var.get()}-{start_week}", f"{self.end_year_var.get()}-{end_week}"
        
        # 默认返回值，防止返回 None
        return "", ""
    
    def set_date_range(self, range_type):
        """设置日期范围"""
        today = datetime.now()
        
        # 初始化 start_date 和 end_date
        start_date = today
        end_date = today
        
        if range_type == "today":
            # start_date 和 end_date 已经初始化为 today
            pass
        elif range_type == "week":
            # 本周开始（周一）
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
        elif range_type == "month":
            # 本月开始
            start_date = today.replace(day=1)
            # 本月结束
            _, last_day = calendar.monthrange(today.year, today.month)
            end_date = today.replace(day=last_day)
        elif range_type == "quarter":
            # 本季度开始
            quarter = (today.month - 1) // 3 + 1
            start_month = 3 * quarter - 2
            start_date = today.replace(month=start_month, day=1)
            # 本季度结束
            end_month = 3 * quarter
            _, last_day = calendar.monthrange(today.year, end_month)
            end_date = today.replace(month=end_month, day=last_day)
        elif range_type == "year":
            # 今年开始
            start_date = today.replace(month=1, day=1)
            # 今年结束
            end_date = today.replace(month=12, day=31)
        
        # 更新控件
        granularity = self.granularity_var.get()
        
        if granularity == "day":
            self.start_date_entry.set_date(start_date)
            self.end_date_entry.set_date(end_date)
            
        elif granularity == "month":
            # 设置年月选择器
            self.start_year_var.set(str(start_date.year))
            self.start_month_var.set(f"{start_date.month:02d}")
            self.end_year_var.set(str(end_date.year))
            self.end_month_var.set(f"{end_date.month:02d}")
            
        elif granularity == "year":
            self.start_year_var.set(str(start_date.year))
            self.end_year_var.set(str(end_date.year))
            
        elif granularity == "quarter":
            start_quarter = (start_date.month - 1) // 3 + 1
            end_quarter = (end_date.month - 1) // 3 + 1
            self.start_year_var.set(str(start_date.year))
            self.start_quarter_var.set(f"{start_quarter}")
            self.end_year_var.set(str(end_date.year))
            self.end_quarter_var.set(f"{end_quarter}")
            
        elif granularity == "week":
            # 计算ISO周数
            start_iso_year, start_iso_week, _ = start_date.isocalendar()
            end_iso_year, end_iso_week, _ = end_date.isocalendar()
            self.start_year_var.set(str(start_iso_year))
            self.start_week_var.set(str(start_iso_week).zfill(2))
            self.end_year_var.set(str(end_iso_year))
            self.end_week_var.set(str(end_iso_week).zfill(2))
    
    def export_data(self):
        """导出数据"""
        # 获取输入值
        api_base_url = self.api_base_url_var.get()
        id_param = self.id_entry.get()
        backup_folder = self.backup_folder_var.get()
        upload_folder = self.upload_folder_var.get()
        file_format = self.format_var.get().lower()
        
        # 获取时间值
        try:
            start_time, end_time = self.get_time_values()
        except Exception as e:
            messagebox.showerror("错误", f"获取时间值时出错: {str(e)}")
            logging.error(f"获取时间值时出错: {str(e)}")
            return
            
        # 验证输入
        if not api_base_url:
            messagebox.showerror("错误", "请输入API地址")
            return
            
        if not id_param:
            messagebox.showerror("错误", "请输入ID")
            return
            
        if not start_time or not end_time:
            messagebox.showerror("错误", "请选择开始时间和结束时间")
            return
            
        if not backup_folder:
            messagebox.showerror("错误", "请选择备份文件夹")
            return
            
        if not upload_folder:
            messagebox.showerror("错误", "请选择上传文件夹")
            return
            
        # 生成文件名
        filename = self.generate_filename()
        backup_path = os.path.join(backup_folder, filename)
        upload_path = os.path.join(upload_folder, filename)
        
        # 更新状态
        self.status_label.config(text="正在获取数据...", foreground="blue")
        self.root.update()
        
        logging.info(f"开始导出数据: API={api_base_url}, ID={id_param}, 时间范围={start_time} 到 {end_time}")
        
        try:
            # 从API获取数据
            json_data = DataExporter.fetch_data_from_api(api_base_url, id_param, start_time, end_time)
            
            if not json_data:
                self.status_label.config(text="获取数据失败", foreground="red")
                messagebox.showerror("错误", "无法从API获取数据")
                return
                
            # 解析API响应
            cellset = DataExporter.parse_api_response(json_data)
            
            if not cellset:
                self.status_label.config(text="解析数据失败", foreground="red")
                messagebox.showerror("错误", "无法解析API返回的数据")
                return
                
            # 导出数据
            self.status_label.config(text="正在导出数据...", foreground="blue")
            self.root.update()
            
            success = False
            if file_format == "excel":
                # 使用pandas导出Excel
                success_backup = DataExporter.export_to_excel_with_pandas(cellset, backup_path)
                success_upload = DataExporter.export_to_excel_with_pandas(cellset, upload_path)
                success = success_backup and success_upload
            elif file_format == "csv":
                success_backup = DataExporter.export_to_csv(cellset, backup_path)
                success_upload = DataExporter.export_to_csv(cellset, upload_path)
                success = success_backup and success_upload
            
            if success:
                self.status_label.config(text=f"导出成功! 文件已保存到: {backup_path} 和 {upload_path}", foreground="green")
                logging.info(f"导出成功: {backup_path} 和 {upload_path}")
            else:
                self.status_label.config(text="导出失败", foreground="red")
                logging.error("导出数据时出错")
                
        except Exception as e:
            self.status_label.config(text=f"错误: {str(e)}", foreground="red")
            messagebox.showerror("错误", f"导出数据时发生错误:\n{str(e)}")
            logging.error(f"导出数据时发生错误: {str(e)}")
    
    def toggle_scheduler(self):
        """切换定时任务状态"""
        if self.scheduler_running:
            self.stop_scheduler()
        else:
            self.start_scheduler()
    
    def start_scheduler(self):
        """启动定时任务"""
        # 验证必要的设置
        if not self.backup_folder_var.get():
            messagebox.showerror("错误", "请先选择备份文件夹")
            return
        if not self.upload_folder_var.get():
            messagebox.showerror("错误", "请先选择上传文件夹")
            return
            
        # 配置定时任务
        schedule_type = self.schedule_type_var.get()
        
        try:
            if schedule_type == "每天":
                # 每天特定时间执行
                schedule_time = self.schedule_time_var.get()
                hours, minutes = map(int, schedule_time.split(':'))
                schedule.every().day.at(f"{hours:02d}:{minutes:02d}").do(self.scheduled_export)
                logging.info(f"定时任务已设置: 每天 {hours:02d}:{minutes:02d} 执行")
                
            elif schedule_type == "每周":
                # 每周特定星期几和时间执行
                schedule_day = self.schedule_day_var.get()
                schedule_time = self.schedule_time_var.get()
                hours, minutes = map(int, schedule_time.split(':'))
                
                # 映射中文星期几到英文
                day_mapping = {
                    "星期一": "monday",
                    "星期二": "tuesday",
                    "星期三": "wednesday",
                    "星期四": "thursday",
                    "星期五": "friday",
                    "星期六": "saturday",
                    "星期日": "sunday"
                }
                
                # 映射星期几到schedule的方法
                day_methods = {
                    'monday': schedule.every().monday,
                    'tuesday': schedule.every().tuesday,
                    'wednesday': schedule.every().wednesday,
                    'thursday': schedule.every().thursday,
                    'friday': schedule.every().friday,
                    'saturday': schedule.every().saturday,
                    'sunday': schedule.every().sunday
                }
                
                english_day = day_mapping.get(schedule_day)
                if english_day and english_day in day_methods:
                    day_methods[english_day].at(f"{hours:02d}:{minutes:02d}").do(self.scheduled_export)
                    logging.info(f"定时任务已设置: 每周 {schedule_day} {hours:02d}:{minutes:02d} 执行")
                else:
                    messagebox.showerror("错误", "无效的星期几设置")
                    return
                    
            elif schedule_type == "每月":
                # 每月特定日期和时间执行
                schedule_day = int(self.schedule_day_var.get())
                schedule_time = self.schedule_time_var.get()
                hours, minutes = map(int, schedule_time.split(':'))
                
                # 使用lambda包装，因为schedule不支持直接设置每月某日
                def monthly_job():
                    now = datetime.now()
                    if now.day == schedule_day:
                        self.scheduled_export()
                
                # 每天检查是否是指定日期
                schedule.every().day.at(f"{hours:02d}:{minutes:02d}").do(monthly_job)
                logging.info(f"定时任务已设置: 每月 {schedule_day}日 {hours:02d}:{minutes:02d} 执行")
                
            elif schedule_type == "每小时":
                # 每小时特定分钟执行
                schedule_minute = int(self.schedule_minute_var.get())
                
                def hourly_job():
                    now = datetime.now()
                    if now.minute == schedule_minute:
                        self.scheduled_export()
                
                # 每分钟检查是否是指定分钟
                schedule.every().minute.do(hourly_job)
                logging.info(f"定时任务已设置: 每小时 {schedule_minute}分 执行")
            
            # 启动定时任务线程
            self.scheduler_running = True
            self.scheduler_thread = threading.Thread(target=self.run_scheduler, daemon=True)
            self.scheduler_thread.start()
            
            self.schedule_button.config(text="停止定时任务")
            self.status_label.config(text=f"定时任务已启动: {schedule_type}", foreground="blue")
            
        except Exception as e:
            messagebox.showerror("错误", f"配置定时任务时出错: {str(e)}")
            logging.error(f"配置定时任务时出错: {str(e)}")
    
    def stop_scheduler(self):
        """停止定时任务"""
        self.scheduler_running = False
        schedule.clear()
        self.schedule_button.config(text="启动定时任务")
        self.status_label.config(text="定时任务已停止", foreground="green")
        logging.info("定时任务已停止")
    
    def run_scheduler(self):
        """运行定时任务循环"""
        while self.scheduler_running:
            schedule.run_pending()
            time.sleep(1)
    
    def scheduled_export(self):
        """定时执行导出任务"""
        # 在主线程中执行GUI操作
        self.root.after(0, self._scheduled_export_gui)
    
    def _scheduled_export_gui(self):
        """在GUI线程中执行导出任务"""
        try:
            self.status_label.config(text=f"定时任务执行中: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", foreground="blue")
            logging.info("定时任务开始执行导出操作")
            self.export_data()
        except Exception as e:
            self.status_label.config(text=f"定时任务执行失败: {str(e)}", foreground="red")
            logging.error(f"定时任务执行失败: {str(e)}")

# 主程序入口
if __name__ == "__main__":
    try:
        # 添加启动延迟，防止窗口立即关闭
        time.sleep(1)
        
        root = tk.Tk()
        app = DataExporterApp(root)
        
        # 添加启动日志
        logging.info("应用程序启动成功")
        root.mainloop()
    except Exception as e:
        # 记录启动错误
        logging.critical(f"应用程序启动失败: {str(e)}", exc_info=True)
        
        # 创建错误提示窗口
        error_root = tk.Tk()
        error_root.title("启动错误")
        error_root.geometry("400x200")
        
        ttk.Label(error_root, text="应用程序启动失败", font=("Arial", 14, "bold"), foreground="red").pack(pady=20)
        ttk.Label(error_root, text=f"错误信息: {str(e)}").pack(pady=10)
        ttk.Label(error_root, text="请查看日志文件获取详细信息").pack(pady=10)
        
        # 添加日志文件位置提示
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        log_file = os.path.join(base_dir, "data_export.log")
        ttk.Label(error_root, text=f"日志文件位置: {log_file}").pack(pady=10)
        
        error_root.mainloop()