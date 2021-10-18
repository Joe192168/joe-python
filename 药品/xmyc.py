import re
import os
import time
import datetime
import random
import openpyxl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait

# 获取当前时间
nowTime = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
# 设置全局变量用来存储数据
data_list = []

# 无界面 测试有问题暂时有问题
chrome_options = webdriver.ChromeOptions()
# 非沙盒模式运行
chrome_options.add_argument('--no-sandbox')
# headless模式启动
# chrome_options.add_argument('--headless')
# 谷歌文档提到需要加上这个属性来规避bug
chrome_options.add_argument('--disable-gpu')
# 最大化
chrome_options.add_argument("--start-maximized")
# 不加载图片
chrome_options.add_argument('blink-settings=imagesEnabled=false')
chromeDriverPath = r'..\tools\chromedriver.exe'
browser = webdriver.Chrome(executable_path=chromeDriverPath, options=chrome_options)
# 设置等待时间
wait = WebDriverWait(browser, 50)

# 登陆
def login():
    print("*************爬取数据操作界面***************")
    url = input("如：www.baidu.com 请输入网址：")
    username = input("用户名：")
    password = input("登录密码：")

    browser.get("https://" + "www.xmyc.com.cn/login")
    browser.find_element_by_xpath('//input[@placeholder="请输入用户名/手机号"]').send_keys("13679255320")
    browser.find_element_by_xpath('//input[@placeholder="密码"]').send_keys("123456")
    browser.find_element_by_xpath('//button[@class="ivu-btn ivu-btn-success ivu-btn-long ivu-btn-large"]').click()

# 杀死浏览器进程
def kill_driver():
    # 杀死这个chromedriver进程，因为每次启动都会打开，所以需要kill，这里用的chrome浏览器
    os.system('chcp 65001')
    os.system("taskkill /f /im chromedriver.exe")

# 新建excel
def creatwb(wbname):
    wb = openpyxl.Workbook()
    wb.save(filename=wbname)
    print("新建Excel："+wbname+"成功")

# 搜索
def search(keyword):
    try:
        # 等到搜索框加载出来
        submit = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//button[@class="search-btn ivu-btn ivu-btn-primary"]'))
        )
        flag1 = isElementExist(browser, "close shadow")
        if flag1:
            browser.find_element_by_xpath('//div[@class="close shadow"]').click()
        # 等到搜索按钮可以被点击
        browser.find_element_by_xpath('//input[@placeholder="请输入药品名/助记码/厂家/商品货号"]').send_keys(keyword)
        # 点击
        submit.click()
        # 设置随机延迟
        time.sleep(random.randint(1, 3))
        # 总页数
        totalPages = 0
        # 判断是否搜索到药品信息
        flag2 = isElementExist(browser, "goods-item")
        if not flag2:
            print("没有搜索到该药品数据，关闭当前程序重新进行搜索！")
        else:
            # 模拟用户点击
            #browser.find_element_by_xpath('//input[@class="ivu-checkbox-input"]').click()
            # 判断是否有分页元素存在
            total = wait.until(
                EC.presence_of_element_located((By.XPATH, './/div[@class="sort-bar__right"]//span'))
            )
            total = int(re.sub("\D", "", total.text))
            # 一页显示条数
            row = 30
            # 总页数
            totalPages = total / row
            if total % row != 0:
                totalPages = int(totalPages+1)

            print('总共页数：%s' % totalPages)
            # 滑动到底部，加载出商品信息
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            wait.until(
                EC.presence_of_all_elements_located((By.XPATH, './/ul[@class="footer-page ivu-page"]'))
            )
            print("第", 1, "页：")
            # 设置计时器
            download_web(3)
            # 调用提取数据的函数
            prase_html()
        # 返回总页数
        return totalPages
    except TimeoutError:
        print("请求超时，重新搜索该药品！。。。。。。。。")
        search()
    except Exception as e:
        print(e)
        # 关闭浏览器
        kill_driver()

# 页面加载计时器
def download_web(c_time):
    # 设置随机延迟
    for k in range(c_time,-1,-1):
        print('\r', '距离页面数据加载结束还有 %s 秒！' % str(k).zfill(2), end='')
        time.sleep(1)
    print('\r', '{:^20}'.format('页面加载结束！'))

# 判断元素是否存在方法
def isElementExist(browser,element):
    try:
        browser.find_element_by_xpath('//div[@class="'+element+'"]')
        return True
    except:
        return False

# 提取页面数据
def prase_html():
    # 开始提取信息,找到ul标签下的全部li标签
    try:
        lis = browser.find_elements_by_class_name('goods-item')
        # 遍历
        for pl in lis:
            # 名称
            title = pl.find_element_by_xpath('.//div[@class="name"]').text
            # 售价
            price = pl.find_element_by_xpath('.//div[@class="price"]//span[@class="goods-sale-price"]').text
            # 规格
            specs = pl.find_element_by_xpath('.//div[@class="goods-info"]//a[@class="tit"]//p').text
            # 有效期 库存 件装量 中包装
            ykjz = pl.find_elements_by_xpath('.//div[@class="goods-info"]//a[@class="tit"]//p[@class="time-stock"]')
            # 厂家
            li_caption = pl.find_element_by_xpath('.//div[@class="goods-info"]//a[@class="tit"]//p[@class="yc-name"]').text
            # 写入字典
            data_dict = []
            data_dict.append(title)
            data_dict.append(price)
            data_dict.append(specs)

            # 有效期 库存
            yk = ykjz[0].text.split("\n")
            data_dict.append(yk[0])
            data_dict.append(yk[1])

            # 件装量 中包装
            jz = ykjz[1].text.split("\n")
            data_dict.append(jz[0])
            data_dict.append(jz[1])

            data_dict.append(li_caption)
            print(data_dict)
            # 写入全局变量
            data_list.append(data_dict)
    except Exception as e:
        print(e)
        kill_driver()

# 分页
def next_page():
    try:
        # 滑动到底部
        scroll_add_crowd_button = browser.find_element_by_xpath('.//div[@class="footer-wrap"]')
        browser.execute_script("arguments[0].scrollIntoView();", scroll_add_crowd_button)
        # 设置随机延迟
        time.sleep(random.randint(1, 3))
        # 翻页按钮
        button = wait.until(
            EC.element_to_be_clickable((By.CLASS_NAME, 'ivu-page-next'))
        )
        # 翻页动作
        button.click()
        wait.until(
            EC.presence_of_all_elements_located((By.XPATH, './/div[@class="goods-warp"]'))
        )
        # 等到商品都加载出来
        # 滑动到底部，加载出后商品信息
        scroll_add_crowd_button = browser.find_element_by_xpath('.//div[@class="footer-wrap"]')
        browser.execute_script("arguments[0].scrollIntoView();", scroll_add_crowd_button)
        wait.until(
            EC.presence_of_all_elements_located((By.XPATH, './/ul[@class="footer-page ivu-page"]'))
        )
        # 等到最后商品都加载出来
        # 判断翻页成功,高亮的按钮数字与设置的页码一样
        # 设置计时器
        download_web(3)
        # 调用提取数据的函数
        prase_html()
    except TimeoutError:
        return next_page()

# 保存页面数据
def write_excel(fileName,d_list):
    wb = Workbook()
    # 写入表头
    dilei_head = ['名称', '售价', '规格', '有效期', '库存', '件装量', '中包装', '厂家']
    sheet0Name = '药品数据信息'
    sheet0 = wb.create_sheet(sheet0Name, index=0)
    for i, item in enumerate(dilei_head):
        sheet0.cell(row=1, column=i+1, value=item.encode('utf-8'))
    # 写入数据
    for i, data in enumerate(d_list):
        for j, item in enumerate(data):
            sheet0.cell(row=i+2, column=j+1, value=item.encode('utf-8'))

    wb.save(fileName)
    print('excel文件写入完成')

# 根据关键字搜索导出exlce
def export_exlce(keyword):
    # 组装exlce文件名称
    _exclName = keyword + nowTime + ".xlsx"
    total = int(search(keyword))
    if total > 0:
        # 初始化exlce文件
        # creatwb(_exclName)
        for i in range(2, total + 1):
            print("第", i, "页：")
            next_page()
        # 保存数据到exlce中
        write_excel(_exclName, data_list)
        # 清空全局数据集合
        data_list.clear()

# 主方法
def main():
    print("*************开始爬取有货药品数据请稍等***************")
    num = input("请选择爬取数据选项数字：1、取文本  2、单品手动输入\n")
    if num == '1':
        if os.path.exists("../药品品种.txt"):
            # 打开文件
            with open("../药品品种.txt", "r", encoding="utf-8") as f:
                for line in f.readlines():
                    # 去掉列表中每一个元素的换行符
                    data = line.strip('\n')
                    print("开始加载品种：", data)
                    # 导出exlce
                    export_exlce(data)
        else:
            print("读取文本文件不存在，请检查！")
    elif num == '2':
        # 关键词
        keyword = input("请输入要搜索药品名称：")
        # 导出exlce
        export_exlce(keyword)
    else:
        print("输入选项不正确，请重新开始输入！")
        main()

    # 关闭浏览器
    browser.close()

if __name__ == "__main__":
    login()
    # 主方法入口
    main()
    input("请按回车键退出！")
    # 关闭浏览器进程
    kill_driver()