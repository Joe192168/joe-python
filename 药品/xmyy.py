import requests
import json
import time
import datetime
from openpyxl import Workbook

# 登陆类
class Duologin(object):
    def __init__(self):
        # 网站header信息
        self.headers = {
            "Content-Type": "application/json; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36 Edg/92.0.902.67"}
        # 登陆url
        self.login_url = "https://mall.xmyc.com.cn/user/login"
        # 会话
        self.session = requests.Session()
        # token信息
        self.token = ""
        # 页码
        self.page_num = 0
        # 页数
        self.pagesize = 0
        # 总页数
        self.total_page = 0
        # 设置全局变量用来存储数据
        self.data_list = []
        # msg请求信息
        self.msg = ""

    def login(self, username, password):
        post_data = {
            "username": username,
            "password": password
        }
        try:
            response = self.session.post(self.login_url, data=json.dumps(post_data), headers=self.headers)
            jwt = json.loads(response.text)
            # 登陆打印
            # print(json.loads(response.text))
            self.msg = json.loads(response.text)
            # 获取token
            self.token = jwt["data"]["shopToken"]["token"]
        except Exception as e:
            print(e)


    # 搜索
    def search(self, keyword):
        self.headers = {
            "Authorization": self.token,
            "Content-Type": "application/json; charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36 Edg/92.0.902.67"}

        url2 = "https://mall.xmyc.com.cn/business/center/search/searchGoods?pageNum=1&pageSize=30&keyWord="+keyword+""
        # 通过搜索获取
        f = self.session.get(url2, headers=self.headers)
        # 搜索打印
        # print(json.loads(f.text.encode("utf-8")))
        data = json.loads(f.text.encode("utf-8"))
        # 页码
        self.page_num = data["data"]["pageInfo"]["pageable"]["pageNum"]
        # 页数
        self.pagesize = data["data"]["pageInfo"]["pageable"]["pageSize"]
        # 总页数
        self.total_page = data["data"]["pageInfo"]["pageable"]["totalPage"]
        # 分页
        for i in range(1, self.total_page+1):
            print("第", i, "页")
            # 页面时间加载器
            Duologin.loadTime(self, 3)
            url2 = "https://mall.xmyc.com.cn/business/center/search/searchGoods?pageNum="+str(i)+"&pageSize="+str(self.pagesize)+"&keyWord="+keyword+""
            # 通过搜索获取
            f = self.session.get(url2, headers=self.headers)
            data = json.loads(f.text.encode("utf-8"))
            goods_list = data["data"]["pageInfo"]["pageable"]["list"]
            self.datalist(goods_list)

    # 根据药品集合组装数据
    def datalist(self, goods_list):
        # 遍历
        for goods in goods_list:
            # 写入字典
            datadict = []
            # 名称
            datadict.append(goods["goodsName"])
            # 售价
            datadict.append(goods["salePrice"])
            # 规格
            datadict.append(goods["attrName"])
            # 有效期
            datadict.append(goods["expireTime"])
            # 库存
            datadict.append(goods["stockNum"])
            # 件装量
            datadict.append(goods["packageNum"])
            # 中包装
            datadict.append(str(goods["middlePackage"])+goods["packageUnit"])
            # 厂家
            datadict.append(goods["productionName"])
            # 写入全局变量
            self.data_list.append(datadict)

    # 保存页面数据
    def write_excel(self, filename):
        # 获取当前时间
        nowTime = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        filename = filename + nowTime + ".xlsx"
        wb = Workbook()
        # 写入表头
        head = ['名称', '售价', '规格', '有效期', '库存', '件装量', '中包装', '厂家']
        sheet0name = '药品数据信息'
        sheet0 = wb.create_sheet(sheet0name, index=0)
        for i, item in enumerate(head):
            sheet0.cell(row=1, column=i+1, value=item.encode('utf-8'))
        # 写入数据
        for i, data in enumerate(self.data_list):
            for j, item in enumerate(data):
                sheet0.cell(row=i+2, column=j+1, value=str(item).encode('utf-8'))

        wb.save(filename)
        print('excel文件写入完成')

    # 页面加载计时器
    def loadTime(self, c_time):
        # 设置随机延迟
        for k in range(c_time,-1,-1):
            print('\r', '距离页面数据加载结束还有 %s 秒！' % str(k).zfill(2), end='')
            time.sleep(1)
        print('\r', '{:^20}'.format('页面数据加载结束！'))

if __name__ == "__main__":
    login = Duologin()
    print("*************开始爬取有货药品数据请稍等***************")
    username = input("请输入登陆账号：\n")
    password = input("请输入登陆密码：\n")
    # 登陆
    login.login(username, password)
    if login.msg["code"] == 200:
        print(login.msg["data"]["msg"])
        keyword = input("请输入要搜索的药品名称：\n")
        # 搜索
        login.search(keyword)
        # 导出excel数据
        login.write_excel(keyword)
    else:
        print(login.msg["message"])