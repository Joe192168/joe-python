import openpyxl
from openpyxl import Workbook

#新建excel
def creatwb(wbname):
    wb=openpyxl.Workbook()
    wb.save(filename=wbname)
    print ("新建Excel："+wbname+"成功")

# 写入excel文件中 date 数据，date是list数据类型， fields 表头
def write_excel(fileName,d_list):
    wb = Workbook()
    #写入表头
    dilei_head = ['药品名称','售价','销量','阶梯满减','满减']
    sheet0Name = '药品数据信息'
    sheet0 = wb.create_sheet(sheet0Name, index=0)
    for i, item in enumerate(dilei_head):
        sheet0.cell(row = 1,column=i+1,value=item.encode('utf-8'))
    #写入数据
    for i, data in enumerate(d_list):
        for j, item in enumerate(data):
            sheet0.cell(row = i+2,column=j+1,value=item.encode('utf-8'))

    wb.save(fileName+'.xlsx')
    print("保存成功")

if __name__ == "__main__":
    creatwb("aa.xlsx")
    data_dict = []#写入字典
    data_dict.append('厄贝沙坦片(安博维片)(禁销省外)【】 150mg*7s 赛诺菲(杭州)制药有限公司')
    data_dict.append('28.92')
    data_dict.append('50607')
    data_dict.append('满减促销：单笔购满500元，总价立减0.5%；满1000元，总价立减1%；满2000元，总价立减2%')
    data_dict.append('在线支付满299元即享2.5%优惠')
    data_dict2 = []#写入字典
    data_dict2.append('厄贝沙坦片')
    data_dict2.append('28.92')
    data_dict2.append('50607')
    data_dict2.append('满减促销：单笔购满500元，总价立减0.5%；满1000元，总价立减1%；满2000元，总价立减2%')
    data_dict2.append('')
    data_list = []
    data_list.insert(0,data_dict)
    data_list.insert(1,data_dict2)
    write_excel("aa",data_list)