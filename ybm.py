import re
import os
import time
import datetime
import random
import openpyxl
import sys
import traceback
from time import sleep
from openpyxl import Workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ActionChains

# 获取当前时间
nowTime = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
# 设置全局变量用来存储数据
data_list = []
# 设置商品品种数组
data_varieties = []

# 无界面 测试有问题暂时有问题
chrome_options = webdriver.ChromeOptions()
# chrome_options.add_argument('--no-sandbox')#非沙盒模式运行
# chrome_options.add_argument('--headless') #headless模式启动
# chrome_options.add_argument('--disable-gpu')# 谷歌文档提到需要加上这个属性来规避bug
chrome_options.add_argument("--start-maximized") #最大化
chrome_options.add_argument('blink-settings=imagesEnabled=false')#不加载图片
chromeDriverPath = r'.\tools\chromedriver.exe'
browser = webdriver.Chrome(executable_path=chromeDriverPath,chrome_options=chrome_options)
# 设置等待时间
wait = WebDriverWait(browser, 50)

# 登陆
def login():
    print("*************要帮忙爬取数据操作界面***************")
    # 用户名
    username = input("用户名：")
    # 登录密码
    password = input("登录密码：")
    # 爬虫网址
    try:
        # 页面窗口最大化
        # browser.maximize_window()
        browser.get('http://www.ybm100.com/login/login.htm')
        browser.find_element_by_xpath('//*[@id="inputPhone"]').send_keys('13368049626')
        browser.find_element_by_xpath('//*[@id="inputPassword"]').send_keys('qfgdyf100')
        browser.find_element_by_xpath('//*[@id="loginForm"]/div[3]/div/button').click()
    except TimeoutException:
        print(u'页面加载超过设定时间，超时')
        # 当页面加载时间超过设定时间，
        browser.close()

# 页面加载计时器
def download_web(c_time):
    # 设置随机延迟
    for k in range(c_time,-1,-1):
        print('\r','页面加载倒计时，还有 %s 秒！请稍等。。。' % str(k).zfill(2),end='')
        time.sleep(1)
    print('\r','{:^20}'.format('页面加载完毕！'))

# 搜索
def search(keyword):
    try:
        # 主页面句柄  每个浏览器标签页都有一个句柄
        mainhandle = browser.current_window_handle
        # 等到搜索框加载出来
        submit = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="search-btn"]'))
        )
        # 在搜索之前清空搜索框内容
        browser.find_element_by_xpath('//*[@id="search"]').clear()
        # 等到搜索按钮可以被点击
        browser.find_element_by_xpath('//*[@id="search"]').send_keys(keyword)
        submit.click()#点击
        # 获取当前窗口句柄
        handles = browser.window_handles
        # 轮流得出标签页的句柄 切换窗口
        for handle in handles:
            if handle != mainhandle:
                browser.close() # 关闭第一个窗口
                browser.switch_to.window(handle)
        # 判断是否搜索到药品信息
        div_text = browser.find_elements_by_xpath('.//div[@class="main"]/div')
        if div_text[4].text.find("抱歉，没有找到商品")>=0:
            print("没有搜索到该药品数据，重新进行搜索！")
            # 重新开始搜索
            main()
        # browser.find_element_by_link_text("只看有货").click()#模拟用户点击
        # 通过css 找到type = checkbox
        # browser.find_element_by_name('hasStock').click()
        # 判断是否有分页元素存在
        falg2 = isElementExist(browser,".page")
        if falg2:
            total = wait.until(
                EC.presence_of_element_located((By.XPATH, './/.//div[@class="page"]/div/div/span'))
            )
            total = re.sub("\D", "", total.text)
            print("总共页数："+total)
        else:
            print("总共页数：1")
            # 默认只有一页
            total = 1
        # 滑动到底部，加载出商品信息
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # wait.until(
        #     EC.presence_of_all_elements_located((By.XPATH, './/div[@class="mrth-new"]'))
        # )
        # 隐藏搜索浮动框
        # js_div = 'document.getElementsByClassName("followSearchPanel")[0].style="display: none;"'
        # browser.execute_script(js_div)
        # 获取网页信息
        html = browser.page_source
        print("第", 1, "页：")
        # 设置计时器
        download_web(3)
        # 总页数转换int
        total = int(total)
        # 调用提取数据的函数
        extract_html(html)
        # 返回总页数
        return total
    except TimeoutError:
        print("请求超时，重新搜索该药品！。。。。。。。。")
        main()
    except Exception as e:
        print(e)
        # 关闭浏览器
        browser.close()

# 分页
def next_page(page_number):
    try:
        # 滑动到底部
        scroll_add_crowd_button = browser.find_element_by_xpath('.//div[@class="page"]')
        browser.execute_script("arguments[0].scrollIntoView();",scroll_add_crowd_button)
        div_text = browser.find_elements_by_xpath('.//div[@class="main"]/div')
        # 设置随机延迟
        time.sleep(random.randint(1, 3))
        # # 隐藏搜索浮动框
        # js_div = 'document.getElementsByClassName("followSearchPanel")[0].style="display: none;"'
        # browser.execute_script(js_div)
        button = wait.until(
            EC.element_to_be_clickable((By.CLASS_NAME, 'next'))
        )
        # 翻页按钮
        button.click()
        # 翻页动作
        wait.until(
            EC.presence_of_all_elements_located((By.XPATH, './/div[@class="main"]//ul/li'))
        )
        # 等到商品都加载出来
        # 滑动到底部，加载出后商品信息
        scroll_add_crowd_button = browser.find_element_by_xpath('.//div[@class="page"]')
        browser.execute_script("arguments[0].scrollIntoView();",scroll_add_crowd_button)
        wait.until(
            EC.presence_of_all_elements_located((By.XPATH, './/div[@class="main"]//ul/li'))
        )
        # 等到最后商品都加载出来
        # 隐藏搜索浮动框
        # js_div = 'document.getElementsByClassName("followSearchPanel")[0].style="display: none;"'
        # browser.execute_script(js_div)
        # wait.until(
        #     EC.text_to_be_present_in_element((By.CLASS_NAME, "active"), str(page_number))
        # )# 判断翻页成功,高亮的按钮数字与设置的页码一样
        # 获取网页信息
        html = browser.page_source
        # 设置计时器
        download_web(3)
        # 调用提取数据的函数
        extract_html(html)
    except TimeoutError:
        return next_page(page_number)

# 提取页面数据
def extract_html(html):
    # 开始提取信息,找到ul标签下的全部li标签
    try:
        ul = browser.find_elements_by_xpath('.//div[@class="main"]//ul[@class="mrth-new clearfix"]/li')
        # 遍历
        for li in ul:
            # 药品名称
            title = li.find_element_by_xpath('.//div[@class="row2"]/a').text
            # 售价
            sp_price = li.find_element_by_xpath('.//div[@class="row3"]')
            # 零售价
            sp_kongxiao = li.find_elements_by_xpath('.//div[@class="row-last"]//div[@class="kongxiao-box"]/span')
            # 毛利
            sp_maoli = li.find_elements_by_xpath('.//div[@class="row-last"]//div[@class="maoli-box"]/span')
            # 药品生成商
            company = li.find_element_by_xpath('.//div[@class="row5 text-overflow"]').text
            # 药品供应商
            gys = li.find_element_by_xpath('.//div[@class="row7"]/a').text
            # 售价
            if sp_price.find_element_by_xpath('//span[@class="price"]'):
                price = sp_price.text
            else:
                price = "价格签署协议可见"
            # 零售价
            if sp_kongxiao:
                retail_price = sp_kongxiao[1].text
            else:
                retail_price = ""
            # 毛利
            if sp_maoli:
                maoli = sp_maoli[1].text
            else:
                maoli = ""
            # 写入字典
            data_dict = []
            data_dict.append(title)
            data_dict.append(price)
            data_dict.append(retail_price)
            data_dict.append(maoli)
            data_dict.append(company)
            data_dict.append(gys)
            print(data_dict)
            # 写入全局变量
            data_list.append(data_dict)
    except Exception as e:
        print('str(Exception):\t', str(Exception))
        print('str(e):\t\t', str(e))
        print('repr(e):\t', repr(e))
        # Get information about the exception that is currently being handled
        exc_type, exc_value, exc_traceback = sys.exc_info()
        print('e.message:\t', exc_value)
        print("Note, object e and exc of Class %s is %s the same." %
              (type(exc_value), ('not', '')[exc_value is e]))
        print('traceback.print_exc(): ', traceback.print_exc())
        print('traceback.format_exc():\n%s' % traceback.format_exc())

# 判断元素是否存在方法
def isElementExist(browser,css):
    try:
        browser.find_element_by_css_selector(css)
        return True
    except:
        return False

# 封装一个函数，用来判断属性值是否存在
def isElementPresent(driver, path):
    # 用来判断元素标签是否存在
    try:
        driver.find_element_by_xpath(path)
    # 原文是except NoSuchElementException, e:
    except NoSuchElementException as e:
        # 发生了NoSuchElementException异常，说明页面中未找到该元素，返回False
        return False
    else:
        # 没有发生异常，表示在页面中找到了该元素，返回True
        return True

# 新建excel
def creat_excel(wbname):
    wb=openpyxl.Workbook()
    wb.save(filename=wbname)
    print("新建Excel："+wbname+"成功")

# 保存页面数据
def write_excel(sheet_name,file_name,d_list):
    wb = Workbook()
    # 写入表头
    table_head = ['药品名称','售价','零售价','毛利','药品生产商','药品供应商']
    # 根据搜索关键字来创建sheet名称
    sheet0_name = sheet_name
    sheet0 = wb.create_sheet(sheet0_name, index=0)
    # 设置表头信息
    for i, item in enumerate(table_head):
        sheet0.cell(row = 1,column=i+1,value=item.encode('utf-8'))
    # 写入数据
    for i, data in enumerate(d_list):
        for j, item in enumerate(data):
            sheet0.cell(row = i+2,column=j+1,value=item.encode('utf-8'))
    # 保存数据
    wb.save(file_name)
    print('excel文件写入完成')

# 杀死浏览器进程
def kill_driver():
    # 杀死这个chromedriver进程，因为每次启动都会打开，所以需要kill，这里用的chrome浏览器
    os.system('chcp 65001')
    os.system("taskkill /f /im chromedriver.exe")

# 获取网站的商品品种总页数
def get_varieties():
    try:
        # 根据a内容进行定位
        all_yp = browser.find_element_by_link_text('全部药品')
        # 单击全药品
        all_yp.click()
        # 获取全药品href，并访问
        browser.get(all_yp.get_property('href'))
        # 定位全部分类
        all_yjlm = browser.find_element_by_id('defaultOnelevelLm')
        # 全部分类span
        all_span = all_yjlm.find_elements_by_xpath('./span')
        # 点击全部分类span单击事件
        all_span[0].click()
        # 判断是否有分页元素存在
        falg2 = isElementExist(browser,".page")
        if falg2:
            total = wait.until(
                EC.presence_of_element_located((By.XPATH, './/.//div[@class="page"]/div/div/span'))
            )
            total = re.sub("\D", "", total.text)
            print("总共页数："+total)
        else:
            print("总共页数：1")
            # 默认只有一页
            total = 1
        print("第", 1, "页：")
        # 设置计时器
        download_web(3)
        # 总页数转换int
        total = int(total)
        # 调用提取数据的函数
        extract_html(browser.page_source)
        # 返回总页数
        return total
    except TimeoutError:
        print("请求超时，重新搜索该药品！。。。。。。。。")
        get_varieties()
    except Exception as e:
        print(e)
        # 关闭浏览器
        browser.close()

# 根据关键字搜索导出exlce
def export_exlce(keyword):
    # 根据搜索名称调用查询方法，并返回总页数
    total = search(keyword)
    # 组装exlce文件名称
    _exclName = keyword + "_" + nowTime + ".xlsx"
    # for i in range(2, 5):
    for i in range(2, total + 1):
        print("第", i, "页：")
        next_page(i)
    # 判断总页数大于0 并创建exlce
    if total > 0:
        # 初始化exlce文件
        creat_excel(_exclName)
    # 保存数据到exlce中
    write_excel(keyword,_exclName,data_list)
    # 清空全局数据集合
    data_list.clear()

# 根据全部品种导出exlce
def all_export_exlce():
    # 根据搜索名称调用查询方法，并返回总页数
    total = get_varieties()
    # 组装exlce文件名称
    _exclName = '全部药品' + "_" + nowTime + ".xlsx"
    # for i in range(2, 5):
    for i in range(2, total + 1):
        print("第", i, "页：")
        next_page(i)
    # 判断总页数大于0 并创建exlce
    if total > 0:
        # 初始化exlce文件
        creat_excel(_exclName)
    # 保存数据到exlce中
    write_excel('全部药品',_exclName,data_list)

# 主方法
def main():
    print("*************开始爬取有货药品数据请稍等***************")
    num = input("请选择爬取数据选项数字：1、取文本  2、自动批量  3、单品手动输入：")
    if num == '3':
        # 关键词
        keyword = input("请输入要搜索药品名称：")
        # 导出exlce
        export_exlce(keyword)
    elif num == '2':
        # 获取所有商品品种
        all_export_exlce()
    elif num == '1':
        if os.path.exists("药品品种.txt"):
            # 打开文件
            with open("药品品种.txt", "r", encoding="utf-8") as f:
                for line in f.readlines():
                    # 去掉列表中每一个元素的换行符
                    data = line.strip('\n')
                    print("开始加载品种：",data)
                    # 导出exlce
                    export_exlce(data)

        else:
            print("读取文本文件不存在，请检查！")
    else:
        print("输入选项不正确，请重新开始输入！")
        main()

    # 关闭浏览器
    browser.close()

if __name__ == "__main__":
    # 登陆
    login()
    # 主方法入口
    main()
    input("请按回车键退出！")
    # 关闭浏览器进程
    kill_driver()