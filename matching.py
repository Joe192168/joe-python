#-*- coding:utf-8 -*-
import re
import time
import datetime
import xlrd,xlwt
import logging
from openpyxl.styles import Font
from openpyxl import Workbook
import Levenshtein

#log日志
logging.basicConfig(filename='log.txt', level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')

#全局字典
data_list = []

#写入excel文件
def write_excel(file_name,d_list):
    try:
        wb = Workbook()
        #sheet名称
        sheet = wb.create_sheet("匹配结果",index=0)#新建一个excel，sheet表
        # 写入表头
        fields = [u'品名A',u'产地A',u'规格A',u'批文A',u'品名B',u'产地B',u'规格B',u'批文B',u'商品编号B',u'价格']
        # 设置表头信息
        field=1
        for field in range(1,len(fields)+1):   # 写入表头
            sheet.cell(row=1,column=field,value=str(fields[field-1])).font = Font(name='微软雅黑', size=12, italic=False, color='FF0004', bold=False)
        # 写入数据
        row1=1
        col1=0
        for row1 in range(2,len(d_list)+2):  # 写入数据
            for col1 in range(1,len(d_list[row1-2])+1):
                sheet.cell(row=row1,column=col1,value=str(d_list[row1-2][col1-1]))
        # 保存数据
        wb.save(file_name)
        print('匹配完成，并保存完毕')
    except:
        logging.exception('write_excel exception：')

def checkText(file_dir,kh_cp,kh_cd,kh_gg,kh_pw,mc_xsd,cd_xsd,gg_xsd):
    try:
        wb = xlrd.open_workbook(file_dir) #打开excel表
        #获取所有sheet工作簿名称
        sheets = wb.sheet_names()
        for item in sheets:
            #通过名字获取某个sheet页的值
            sheet = wb.sheet_by_name(item)
            #获取行数
            nrows = sheet.nrows
            #获取总列数
            #ncols = sheet.ncols
            #print("The sum rows:%d" %nrows)
            #print("The sum cols:%d" %ncols)

            #获取列数
            for i in range(1,nrows):
                # 从第二行开始匹配
                #产品名称
                mb_cp = str(sheet.cell_value(i,0)).strip();
                #产地名称
                mb_cd = str(sheet.cell_value(i,1)).strip();
                #规格
                mb_gg = str(sheet.cell_value(i,2)).strip();
                #批文
                mb_pw = str(sheet.cell_value(i,3)).strip();
                #商品编号
                mb_bh = str(sheet.cell_value(i,4)).strip();
                #价格
                mb_jg = str(sheet.cell_value(i,5)).strip();
                #产品名称
                a = acquaintance(mb_cp,kh_cp)
                if a >= float(mc_xsd):
                    ret_mc = True
                else:
                    ret_mc = False
                #产地
                b = acquaintance(mb_cd,kh_cd)
                if b >= float(cd_xsd):
                    ret_cd = True
                else:
                    ret_cd = False
                #规格
                c = acquaintance(mb_gg.lower(),kh_gg.lower())
                if c >= float(gg_xsd):
                    ret_gg = True
                else:
                    ret_gg = False
                #判断批文是否非空
                if kh_pw:
                    #批文
                    d = KMP_algorithm(mb_pw.lower(),kh_pw.lower())
                    if len(mb_pw)>0 and len(kh_pw)>0 and d >= 0:
                        ret_pw = True
                    else:
                        ret_pw = False
                    if ret_pw and ret_mc and ret_cd and ret_gg:
                        print('根据相似度匹配成功的品名：'+mb_cp+'\t产地：'+mb_cd+'\t规格：'+mb_gg+'\t批准文号：'+mb_pw+'\t商品编号：'+mb_bh+'\t价格：'+mb_jg)
                        data_dict = []
                        data_dict.append(kh_cp)
                        data_dict.append(kh_cd)
                        data_dict.append(kh_gg)
                        data_dict.append(kh_pw)
                        data_dict.append(mb_cp)
                        data_dict.append(mb_cd)
                        data_dict.append(mb_gg)
                        data_dict.append(mb_pw)
                        data_dict.append(mb_bh)
                        data_dict.append(mb_jg)
                        data_list.append(data_dict)
                        ret = False
                        break
                    else:
                        ret = True
                        continue
                else:
                    if  ret_mc and ret_cd and ret_gg:
                        print('根据相似度匹配成功的品名：'+mb_cp+'\t产地：'+mb_cd+'\t规格：'+mb_gg+'\t批准文号：'+mb_pw+'\t商品编号：'+mb_bh+'\t价格：'+mb_jg)
                        data_dict = []
                        data_dict.append(kh_cp)
                        data_dict.append(kh_cd)
                        data_dict.append(kh_gg)
                        data_dict.append(kh_pw)
                        data_dict.append(mb_cp)
                        data_dict.append(mb_cd)
                        data_dict.append(mb_gg)
                        data_dict.append(mb_pw)
                        data_dict.append(mb_bh)
                        data_dict.append(mb_jg)
                        data_list.append(data_dict)
                        ret = False
                        break
                    else:
                        ret = True
                        continue
            return ret

    except:
        logging.exception('checkText exception：')

#根据文件夹 截取文件名称
def getFileList(path_a,path_b,mc_xsd,cd_xsd,gg_xsd):
    try:
        #获取excel文件内容，并判断是否包含
        wb = xlrd.open_workbook(path_a) #打开excel表
        #获取所有sheet工作簿名称
        sheets = wb.sheet_names()
        for item in sheets:
            #通过名字获取某个sheet页的值
            sheet = wb.sheet_by_name(item)
            #获取行数
            nrows = sheet.nrows
            for i in range(1,nrows):
                #产品名称
                cp = str(sheet.cell_value(i,0)).strip();
                #产地名称
                cd = str(sheet.cell_value(i,1)).strip();
                #规格
                gg = str(sheet.cell_value(i,2)).strip();
                #批文
                pw = str(sheet.cell_value(i,3)).strip();
                falg = checkText(path_b,cp,cd,gg,pw,mc_xsd,cd_xsd,gg_xsd)
                if falg:
                    data_dict = []
                    data_dict.append(cp)
                    data_dict.append(cd)
                    data_dict.append(gg)
                    data_dict.append(pw)
                    data_dict.append('')
                    data_dict.append('')
                    data_dict.append('')
                    data_dict.append('')
                    data_dict.append('')
                    data_dict.append('')
                    data_list.append(data_dict)
    except:
        logging.exception('getFileList exception：')

#计算相似度算法
def acquaintance(a,b):
    return Levenshtein.ratio(a,b)

def KMP_algorithm(string, substring):
    '''
    KMP字符串匹配的主函数
    若存在字串返回字串在字符串中开始的位置下标，或者返回-1
    '''
    pnext = gen_pnext(substring)
    n = len(string)
    m = len(substring)
    i, j = 0, 0
    while (i<n) and (j<m):
        if (string[i]==substring[j]):
            i += 1
            j += 1
        elif (j!=0):
            j = pnext[j-1]
        else:
            i += 1
    if (j == m):
        return i-j
    else:
        return -1

def gen_pnext(substring):
    """
    构造临时数组pnext
    """
    index, m = 0, len(substring)
    pnext = [0]*m
    i = 1
    while i < m:
        if (substring[i] == substring[index]):
            pnext[i] = index + 1
            index += 1
            i += 1
        elif (index!=0):
            index = pnext[index-1]
        else:
            pnext[i] = 0
            i += 1
    return pnext

def now():
    return time.strftime('%Y-%m-%d',time.localtime(time.time()))

if __name__ == '__main__':
    #到期时间
    expire = '2020-10-20 00:00:00'
    #根据有效期限进行判断
    if expire < now():
        print('已过期,无法正常使用')
    else:
        print('未过期,可以正常使用')
        mc_xsd = input("品名相似度(如：0-1之间，0.6相当于百分之60的概率)：")
        cd_xsd = input("产地相似度(如：0-1之间，0.6相当于百分之60的概率)：")
        gg_xsd = input("规格相似度(如：0-1之间，0.6相当于百分之60的概率)：")
        #不接受09这样的为整数
        regInt='^0$|^[1-9]\d*$'
        #接受0.00、0.360这样的为小数，不接受00.36，思路:若整数位为零,小数位可为任意整数，但小数位数至少为1位，若整数位为自然数打头，后面可添加任意多个整数，小数位至少1位
        regFloat='^0\.\d+$|^[1-9]\d*\.\d+$'
        regIntOrFloat=regInt+'|'+regFloat#整数或小数
        patternIntOrFloat=re.compile(regIntOrFloat)#创建pattern对象，以便后续可以复用
        if patternIntOrFloat.search(mc_xsd) or patternIntOrFloat.search(cd_xsd) or patternIntOrFloat.search(gg_xsd):
            path_a = input("请输入客户数据的excel文件路径(如：C:/A.xlsx)：")
            path_b = input("请输入目录表的excel文件路径(如：C:/B.xlsx)：")
            start = time.time()
            print("正在匹配，请稍后。。。")
            #匹配数据字典
            getFileList(path_a,path_b,mc_xsd,cd_xsd,gg_xsd)
            # 获取当前时间
            nowTime = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            # 判断字典是否非空
            if data_list:
                #创建excel名称
                excelName = '匹配结果_'+nowTime+'.xlsx';
                print("正在处理数据，请稍等。。。")
                #导出exlce结果
                write_excel(excelName,data_list)
            else:
                print("没有匹配的数据，请重新检查")
            end = time.time()
            print("耗时: %.3fs" % (end - start))
        else:
            print("输入的相似度不是数字，请重新输入")