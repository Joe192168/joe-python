import re
import os
import time
import datetime
import random
import openpyxl
import sys
import traceback
from lxml import etree
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait

#获取当前时间
nowTime = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
#组装exlce文件名称
_exclName = "yp" + nowTime + ".xlsx"
data_list= []#设置全局变量用来存储数据

#无界面 测试有问题暂时有问题
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--no-sandbox')#非沙盒模式运行
#chrome_options.add_argument('--headless') #headless模式启动
chrome_options.add_argument('--disable-gpu')# 谷歌文档提到需要加上这个属性来规避bug
chrome_options.add_argument("--start-maximized") #最大化
chrome_options.add_argument('blink-settings=imagesEnabled=false')#不加载图片
chromeDriverPath = r'.\tools\chromedriver.exe'
browser = webdriver.Chrome(executable_path=chromeDriverPath,chrome_options=chrome_options)
wait =WebDriverWait(browser,50)#设置等待时间

#登陆
def login():
    print("*************爬取数据操作界面***************")
    url = input("如：www.baidu.com 请输入网址：")#爬虫网址
    username = input("用户名：")#用户名
    password = input("登录密码：")#登录密码

    browser.get("http://" + url)
    browser.find_element_by_xpath('//input[@placeholder="用户名"]').send_keys(username)
    browser.find_element_by_xpath('//input[@placeholder="登录密码"]').send_keys(password)
    browser.find_element_by_xpath('//button[@class="loginBtn"]').click()

#页面加载计时器
def download_web(c_time):
    # 设置随机延迟
    for k in range(c_time,-1,-1):
        print('\r','距离页面数据加载结束还有 %s 秒！' % str(k).zfill(2),end='')
        time.sleep(1)
    print('\r','{:^20}'.format('页面加载结束！'))

#搜索
def search(keyword):
    try:
        input = wait.until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "input.search-input"))
        )
        #等到搜索框加载出来
        submit = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='searchForm ']/div/button[1]"))
        )
        #等到搜索按钮可以被点击
        #input[0].send_keys(keyword)#向搜索框内输入关键词
        browser.find_element_by_xpath('//*[@id="searchForm "]/div/input').send_keys(keyword)
        submit.click()#点击
        #判断是否搜索到药品信息
        falg1 = isElementExist(browser,".pl-skin")
        if not falg1:
            print("没有搜索到该药品数据，关闭当前程序重新进行搜索！")
            #关闭浏览器
            browser.close()
        browser.find_element_by_link_text("只看有货").click()#模拟用户点击
        #判断是否有分页元素存在
        falg2 = isElementExist(browser,".pagebar")
        if falg2:
            total = wait.until(
                EC.presence_of_element_located((By.XPATH, './/div[@class="pagebar"]//span'))
            )
            total = re.sub("\D", "", total.text)
            print("总共页数："+total)
        else :
            print("总共页数：1")
            total = 1#默认只有一页
        #滑动到底部，加载出商品信息
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        wait.until(
            EC.presence_of_all_elements_located((By.XPATH, './/div[@class="pl-skin"]'))
        )
        #隐藏搜索浮动框
        js_div = 'document.getElementsByClassName("followSearchPanel")[0].style="display: none;"'
        browser.execute_script(js_div)
        html = browser.page_source#获取网页信息
        print("第", 1, "页：")
        # 设置计时器
        download_web(20)
        prase_html(html)#调用提取数据的函数
        #返回总页数
        return total
    except TimeoutError:
        print("请求超时，重新搜索该药品！。。。。。。。。")
        search()
    except Exception as e:
        print(e)
        #关闭浏览器
        browser.close()

#分页
def next_page(page_number):
    try:
        # 滑动到底部
        scroll_add_crowd_button = browser.find_element_by_xpath('.//div[@class="pages"]')
        browser.execute_script("arguments[0].scrollIntoView();",scroll_add_crowd_button)
        time.sleep(random.randint(1, 3))#设置随机延迟
        #隐藏搜索浮动框
        js_div = 'document.getElementsByClassName("followSearchPanel")[0].style="display: none;"'
        browser.execute_script(js_div)
        button = wait.until(
            EC.element_to_be_clickable((By.CLASS_NAME, 'nextpage'))
        )#翻页按钮
        button.click()# 翻页动作
        wait.until(
            EC.presence_of_all_elements_located((By.XPATH, './/div[@class="pl-skin"]'))
        )#等到商品都加载出来
        # 滑动到底部，加载出后商品信息
        scroll_add_crowd_button = browser.find_element_by_xpath('.//div[@class="pages"]')
        browser.execute_script("arguments[0].scrollIntoView();",scroll_add_crowd_button)
        wait.until(
            EC.presence_of_all_elements_located((By.XPATH, './/div[@class="pl-skin"]'))
        )#等到最后商品都加载出来
        #隐藏搜索浮动框
        js_div = 'document.getElementsByClassName("followSearchPanel")[0].style="display: none;"'
        browser.execute_script(js_div)
        wait.until(
            EC.text_to_be_present_in_element((By.CLASS_NAME, "cur"), str(page_number))
        )# 判断翻页成功,高亮的按钮数字与设置的页码一样
        html = browser.page_source#获取网页信息
        # 设置计时器
        download_web(20)
        #调用提取数据的函数
        prase_html(html)
    except TimeoutError:
        return next_page(page_number)

#提取页面数据
def prase_html(html):
    html = etree.HTML(html)
    # 开始提取信息,找到ul标签下的全部li标签
    try:
        lis = browser.find_elements_by_class_name('pl-skin')
        # 遍历
        for pl in lis:
            is_skus = pl.find_element_by_xpath('.//div')
            #判断有禁售药品或下架
            if is_skus.get_attribute('skus') == 'single':
                # 药品名称
                title = pl.find_element_by_xpath('.//div[@class="p-caption"]//a').text
                # 售价
                price = pl.find_element_by_xpath('.//div[@class="p-priceInfo"]//span[@class="price"]//em')
                # 销量
                sale = pl.find_element_by_xpath('.//.//.//.//div[@class="p-countInfo"]//span[@class="p-sale"]//em')
                # 阶梯满减或满减
                li_caption = pl.find_elements_by_xpath('.//.//.//.//div[@class="promotions-list"]//li[@class="promt_li"]//a')
                if price.text:
                    price = price.text
                else:
                    price = "商家尚未定价"
                if sale:
                    sale = sale.text
                else:
                    sale = ""
                if li_caption:
                    # for labx in li_caption:
                    if li_caption[0].text=="阶梯满减":
                        j_promt_caption = li_caption[1].get_attribute("innerHTML")
                        m_promt_caption = li_caption[3].get_attribute("innerHTML")
                    elif li_caption[0].text=="满减":
                        m_promt_caption = li_caption[1].get_attribute("innerHTML")
                        j_promt_caption = ""
                    elif li_caption[2].text=="满减":
                        j_promt_caption = ""
                        m_promt_caption = li_caption[3].get_attribute("innerHTML")
                    else:
                        j_promt_caption = ""
                        m_promt_caption = ""
                else:
                    j_promt_caption = ""
                    m_promt_caption = ""
                data_dict = []#写入字典
                data_dict.append(title)
                data_dict.append( price)
                data_dict.append(sale)
                data_dict.append(j_promt_caption)
                data_dict.append(m_promt_caption)
                print(data_dict)
                data_list.append(data_dict)#写入全局变量
    except Exception as e:
        print('str(Exception):\t', str(Exception))
        print('str(e):\t\t', str(e))
        print('repr(e):\t', repr(e))
        # Get information about the exception that is currently being handled
        exc_type, exc_value, exc_traceback = sys.exc_info()
        print('e.message:\t', exc_value)
        print("Note, object e and exc of Class %s is %s the same." %
              (type(exc_value), ('not', '')[exc_value is e]))
        print('traceback.print_exc(): ', traceback.print_exc())
        print('traceback.format_exc():\n%s' % traceback.format_exc())

#判断元素是否存在方法
def isElementExist(browser,css):
    try:
        browser.find_element_by_css_selector(css)
        return True
    except:
        return False

#新建excel
def creatwb(wbname):
    wb=openpyxl.Workbook()
    wb.save(filename=wbname)
    print ("新建Excel："+wbname+"成功")

#保存页面数据
def write_excel(fileName,d_list):
    wb = Workbook()
    #写入表头
    dilei_head = ['药品名称','售价','销量','阶梯满减','满减']
    sheet0Name = '药品数据信息'
    sheet0 = wb.create_sheet(sheet0Name, index=0)
    for i, item in enumerate(dilei_head):
        sheet0.cell(row = 1,column=i+1,value=item.encode('utf-8'))
    #写入数据
    for i, data in enumerate(d_list):
        for j, item in enumerate(data):
            sheet0.cell(row = i+2,column=j+1,value=item.encode('utf-8'))

    wb.save(fileName)
    print('excel文件写入完成')

#杀死浏览器进程
def kill_driver():
    #杀死这个chromedriver进程，因为每次启动都会打开，所以需要kill，这里用的chrome浏览器
    os.system('chcp 65001')
    os.system("taskkill /f /im chromedriver.exe")

#主方法
def main():
    #初始化exlce文件
    creatwb(_exclName)
    print("*************开始爬取有货药品数据请稍等***************")
    keyword = input("请输入要搜索药品名称：")#关键词
    total = int(search(keyword))
    # for i in range(2, 5):
    for i in range(2, total + 1):
        print("第", i, "页：")
        next_page(i)
    #保存数据到exlce中
    write_excel(_exclName,data_list)
    #关闭浏览器
    browser.close()

if __name__ == "__main__":
    #key_pass = input("请输入秘钥：")
    #校验秘钥 默认用123 md5加密
    #if key_pass=="202cb962ac59075b964b07152d234b70":
    #登陆
    login()
    #主方法入口
    main()
    # else :
    #     print("秘钥校验不正确，关闭该程序，重新运行！")
    #     #关闭浏览器
    #     browser.close()
    input("请按回车键退出！")
    #关闭浏览器进程
    kill_driver()